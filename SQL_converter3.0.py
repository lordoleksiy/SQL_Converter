import openpyxl
import pyodbc
import time
import traceback
import sys

"""
   Program for transfer data from xlsm file to  database Microsoft SQL Server
   created by Kovtun Oleksiy
"""

err = open("error.txt", "w")
# sys.stderr = error

a = input("Як ввести дані про сервер і таблицю?(з файлу(1)/вручну(2)) ").lower()
if a == "file" or a == "з файлу" or a == "файл" or a == "1":
    with open(r"settings.txt", "r") as f:
        f1 = f.readline()
        f2 = f.readline()
        f3 = f.readline()
        file = f1.split(": ")[1][0:-1]
        server = f2.split(': ')[1][0:-1]
        database = f3.split(": ")[1]
else:
    file = input("Шлях до таблиці: ")
    server = input("Ім'я серверу: ")
    database = input("Ім'я бази данних: ")


start = time.time()
db = pyodbc.connect('Driver={SQL Server};'
                      f'Server={server};'
                      f'Database={database};'
                      'Trusted_Connection=yes;'
                      f'UID={server}')
sql = db.cursor()

print("Зчитування початкових даних...")
book = openpyxl.load_workbook(file, read_only=True)
sheet0 = book['DOV_NT']
sql.execute("Delete From M01")
for row in range(3, sheet0.max_row+1):
    temp = sheet0[row][1].value
    sql.execute(f"Delete From {temp}")
    db.commit()

sql.executemany(f"INSERT INTO M99 VALUES(?, ?, ?, ?, ?, ?, ?);", (list(book["M99"].values))[1:])
db.commit()
print("Сторінка M99 перенесена!")


error = False
for row in range(3, sheet0.max_row+1):
    temp = sheet0[row][1].value
    page = (list(book[sheet0[row][1].value].values))[1:]
    if len(page):
        if temp != "M99":
            if page[0][0]:
                text = ("?, " * len(page[0]))[0:-2] + ");"
                for k in range(len(page)):
                    try:
                        sql.execute(f"INSERT INTO {temp} VALUES(" + text, page[k])
                    except Exception:
                        err.write(f"Помилка виникла на сторінці: {temp}, у рядку: {k+2}\n")
                        err.write(f'Помилка:\n{traceback.format_exc()}')
                        print("\n----------------------------------------------------------------\n\tПомилка!!!!!!!")
                        print(f"Помилка виникла на сторінці: {temp}, у рядку: {k+2}\n")
                        print('Помилка:\n', traceback.format_exc())
                        error = True
                if error:
                    break
                db.commit()
            print(f"Сторінка {temp} перенесена!")
err.close()
db.close()
book.close()
if not error:
    print(f"Програма успішно завершила роботу! Час виконання скрипта: {time.time()-start} c.")
time.sleep(1000)

